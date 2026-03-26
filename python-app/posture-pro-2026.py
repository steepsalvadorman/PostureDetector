"""
Evaluador ROSA — NTP 1173 (INSST, 2022)   v3.1
================================================
INSTALACIÓN:
    pip install opencv-python mediapipe numpy openpyxl pillow
    pip install win10toast   (opcional — notificaciones Windows)
    winsound: incluido en Python/Windows

EMPAQUETAR:
    pip install pyinstaller
    pyinstaller --onefile --windowed --name ROSA_NTP1173 posture-pro-2026.py
"""

import cv2
import mediapipe as mp
import numpy as np
import time, os, sys, threading, queue
import tkinter as tk
from tkinter import ttk, messagebox
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter




try:
    import winsound
    WINSOUND_OK = True
except ImportError:
    WINSOUND_OK = False

try:
    from win10toast import ToastNotifier
    TOAST_OK = True
except ImportError:
    TOAST_OK = False

# ─────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────
INTERVALO_SEG    = 10
GUARDADO_SEG     = 3
MAX_CAMARAS_SCAN = 6
NIVEL_ACCION     = 5   # NTP 1173: score >= 5 requiere actuación inmediata

# ─────────────────────────────────────────────────────────────
# PALETA DE DISEÑO — industrial / oscuro profesional
# ─────────────────────────────────────────────────────────────
C = {
    "bg_deep":   "#0A0C10",
    "bg_panel":  "#111318",
    "bg_card":   "#181C22",
    "bg_border": "#252A34",
    "accent":    "#00C8FF",
    "accent2":   "#0087CC",
    "text_hi":   "#F0F4FF",
    "text_mid":  "#9BAEC8",
    "text_lo":   "#56657A",
    "ok":        "#00D68F",
    "warn":      "#F6AD3B",
    "danger":    "#FF4757",
    "danger2":   "#CC1F2D",
}

NIVELES = {
    1: ("Inapreciable",         "#00D68F"),
    2: ("Bajo",                 "#00B87A"),
    3: ("Medio",                "#F6AD3B"),
    4: ("Medio-Alto",           "#E8823A"),
    5: ("ALTO  —  ACTUAR YA",  "#FF4757"),
}

def nivel_accion(score):
    k = 5 if score >= 5 else max(1, score)
    return NIVELES[k]

# ─────────────────────────────────────────────────────────────
# TABLAS NTP 1173
# ─────────────────────────────────────────────────────────────
# TABLA_A — filas: (A1+A2)-2  |  cols: (A3+A4)-2
# (A1 y A2 mínimo=1 cada uno → suma mínima=2 → índice 0)
TABLA_A = np.array([
    [2,2,3,4,5,6,7,8],[2,2,3,4,5,6,7,8],[3,3,3,4,5,6,7,8],
    [4,4,4,4,5,6,7,8],[5,5,5,5,6,7,8,9],[6,6,6,7,7,8,8,9],[7,7,7,8,8,9,9,9],
])

# TABLA_B — filas: p_B1 (índice directo, fila 0 = "sin teléfono")
#            cols: p_B2 (índice directo, col  0 = "sin pantalla")
# IMPORTANTE: NO restar 1 al indexar. p_B1=1 → fila 1, p_B2=3 → col 3.
TABLA_B = np.array([
    [1,1,1,2,3,4,5,6,6],[1,1,2,2,3,4,5,6,6],[1,2,2,3,3,4,6,7,7],
    [2,2,3,3,4,5,6,8,8],[3,3,4,4,5,6,7,8,8],[4,4,5,5,6,7,8,9,9],[5,5,6,7,8,8,9,9,9],
])

# TABLA_C — filas: p_C1 (índice directo, fila 0 = "sin ratón")
#            cols: p_C2 (índice directo, col  0 = "sin teclado")
# IMPORTANTE: NO restar 1 al indexar. p_C1=1 → fila 1, p_C2=2 → col 2.
TABLA_C = np.array([
    [1,1,1,2,3,4,5,6],[1,1,2,3,4,5,6,7],[1,2,2,3,4,5,6,7],
    [2,3,3,3,5,6,7,8],[3,4,4,5,5,6,7,8],[4,5,5,6,6,7,8,9],
    [5,6,6,7,7,8,8,9],[6,7,7,8,8,9,9,9],
])

# TABLA_D — filas: score_B-1  |  cols: score_C-1
# (scores empiezan en 1 → índice 0 = score 1)
TABLA_D = np.array([
    [1,2,3,4,5,6,7,8,9],[2,2,3,4,5,6,7,8,9],[3,3,3,4,5,6,7,8,9],
    [4,4,4,4,5,6,7,8,9],[5,5,5,5,5,6,7,8,9],[6,6,6,6,6,6,7,8,9],
    [7,7,7,7,7,7,7,8,9],[8,8,8,8,8,8,8,8,9],[9,9,9,9,9,9,9,9,9],
])

def tlu(tabla, fi, ci):
    return int(tabla[int(np.clip(fi, 0, tabla.shape[0]-1))]
                    [int(np.clip(ci, 0, tabla.shape[1]-1))])

def factor_tiempo_F(horas):
    return +1 if horas > 4 else (-1 if horas < 1 else 0)

def calcular_angulo(a, b, c):
    a, b, c = np.array(a), np.array(b), np.array(c)
    ba, bc = a - b, c - b
    cos = np.dot(ba, bc) / (np.linalg.norm(ba) * np.linalg.norm(bc) + 1e-6)
    return float(np.degrees(np.arccos(np.clip(cos, -1, 1))))

def calcular_ROSA(ang_tronco, desv_cuello, ang_rodilla, ang_codo, factor_t):
    # ── SILLA (Tablas A-1 … A-4) ──────────────────────────────
    p_A1 = 1 if 85 <= ang_rodilla <= 100 else 2
    p_A2 = 1
    p_A3 = 1 if 80 <= ang_codo <= 105 else 2
    p_A4 = 1 if 95 <= (90 + ang_tronco) <= 110 else 2

    score_A = tlu(TABLA_A, (p_A1 + p_A2) - 2, (p_A3 + p_A4) - 2)
    score_s = int(np.clip(score_A + factor_t, 1, 10))

    # ── PERIFÉRICOS ────────────────────────────────────────────
    # Tabla B-1: teléfono
    p_B1 = 1   # postura neutra (cuello recto, uso con una mano)

    # Tabla B-2: pantalla — desviación postural del cuello
    p_B2 = 1 if desv_cuello < 10 else (2 if desv_cuello < 20 else 3)

    # FIX v3.1: indexar con p_B1 y p_B2 directamente (sin -1).
    # TABLA_B incluye fila/col 0 para "elemento no presente"; p_Bx=1 → índice 1.
    score_B = tlu(TABLA_B, p_B1, p_B2)

    # Tabla C-1: ratón / Tabla C-2: teclado
    p_C1 = 1 if 80 <= ang_codo <= 110 else 2
    p_C2 = 1 if 80 <= ang_codo <= 110 else 2

    # FIX v3.1: mismo criterio que TABLA_B.
    score_C = tlu(TABLA_C, p_C1, p_C2)

    # Tabla D: periféricos combinados (scores 1-9 → índices 0-8)
    score_D = tlu(TABLA_D, score_B - 1, score_C - 1)

    # ── ROSA FINAL (Tabla E = max) ─────────────────────────────
    rosa = int(np.clip(max(score_s, score_D), 1, 10))

    det = {
        "p_A1_altura":      p_A1,
        "p_A2_profundidad": p_A2,
        "p_A3_reposabrazos":p_A3,
        "p_A4_respaldo":    p_A4,
        "tabla_A":          score_A,
        "factor_F":         factor_t,
        "score_silla":      score_s,
        "p_B1_telefono":    p_B1,
        "p_B2_pantalla":    p_B2,
        "tabla_B":          score_B,
        "p_C1_raton":       p_C1,
        "p_C2_teclado":     p_C2,
        "tabla_C":          score_C,
        "tabla_D":          score_D,
        "rosa_final":       rosa,
    }
    return rosa, score_s, score_D, det

# ─────────────────────────────────────────────────────────────
# AUDIO + NOTIFICACIONES
# ─────────────────────────────────────────────────────────────
def disparar_sonido(veces=4, freq=1200, dur=320):
    def _run():
        for _ in range(veces):
            if WINSOUND_OK:
                try: winsound.Beep(freq, dur); time.sleep(0.08)
                except: pass
            else:
                sys.stdout.write("\a"); sys.stdout.flush(); time.sleep(0.3)
    threading.Thread(target=_run, daemon=True).start()

def disparar_toast(toaster, titulo, cuerpo):
    if toaster and TOAST_OK:
        def _go():
            try: toaster.show_toast(titulo, cuerpo, duration=8, threaded=False)
            except: pass
        threading.Thread(target=_go, daemon=True).start()

# ─────────────────────────────────────────────────────────────
# DETECCIÓN DE CÁMARAS
# ─────────────────────────────────────────────────────────────
def detectar_camaras(max_idx=MAX_CAMARAS_SCAN):
    found = []
    for i in range(max_idx):
        try:
            cap = cv2.VideoCapture(i, cv2.CAP_DSHOW)
            if cap.isOpened():
                ret, _ = cap.read()
                if ret: found.append((i, f"Cámara {i}"))
                cap.release()
        except: pass
    return found or [(0, "Cámara 0 (por defecto)")]

# ─────────────────────────────────────────────────────────────
# EXPORTACIÓN XLSX
# ─────────────────────────────────────────────────────────────
COLS_XLSX = [
    ("id_evaluado",          "id_evaluado"),
    ("nombre_evaluado",      "nombre_evaluado"),
    ("horas_uso_diario",     "horas_uso_diario"),
    ("timestamp",            "timestamp"),
    ("fecha",                "fecha"),
    ("hora",                 "hora"),
    ("rosa_final",           "rosa_final"),
    ("nivel_riesgo",         "nivel_riesgo"),
    ("alerta_disparada",     "alerta_disparada"),
    ("score_silla",          "score_silla"),
    ("score_perifericos_D",  "score_perifericos_D"),
    ("p_A1_altura_asiento",  "p_A1_altura"),
    ("p_A2_profundidad",     "p_A2_profundidad"),
    ("p_A3_reposabrazos",    "p_A3_reposabrazos"),
    ("p_A4_respaldo",        "p_A4_respaldo"),
    ("tabla_A_silla",        "tabla_A"),
    ("factor_tiempo_F",      "factor_F"),
    ("p_B1_telefono",        "p_B1_telefono"),
    ("p_B2_pantalla",        "p_B2_pantalla"),
    ("tabla_B_tel_pant",     "tabla_B"),
    ("p_C1_raton",           "p_C1_raton"),
    ("p_C2_teclado",         "p_C2_teclado"),
    ("tabla_C_rat_tec",      "tabla_C"),
    ("tabla_D_perifericos",  "tabla_D"),
    ("ang_tronco_deg",       "ang_tronco_deg"),
    ("ang_cuello_desv_deg",  "ang_cuello_desv_deg"),
    ("ang_rodilla_deg",      "ang_rodilla_deg"),
    ("ang_codo_deg",         "ang_codo_deg"),
]

def exportar_xlsx(ruta, registros):
    def thin():
        s = Side(style="thin", color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)
    hdrs = [c for c, _ in COLS_XLSX]
    if os.path.exists(ruta):
        wb = load_workbook(ruta); ws = wb.active
        ya = ws.max_row - 1
        nuevos = registros[ya:]
        fi = ws.max_row + 1
    else:
        wb = Workbook(); ws = wb.active; ws.title = "ROSA_Evaluaciones"
        fill_h = PatternFill("solid", fgColor="0D1B2A")
        for ci, col in enumerate(hdrs, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.font      = Font(bold=True, color="00C8FF", size=10, name="Calibri")
            c.fill      = fill_h
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = thin()
        anchos = {
            "timestamp": 17, "fecha": 12, "hora": 10,
            "nombre_evaluado": 22, "id_evaluado": 14,
            "nivel_riesgo": 20, "alerta_disparada": 16,
            "ang_tronco_deg": 15, "ang_cuello_desv_deg": 19,
            "ang_rodilla_deg": 15, "ang_codo_deg": 13,
        }
        for ci, (_, k) in enumerate(COLS_XLSX, 1):
            ws.column_dimensions[get_column_letter(ci)].width = anchos.get(k, 13)
        ws.freeze_panes   = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(hdrs))}1"
        ws.row_dimensions[1].height = 22
        nuevos = registros; fi = 2
    for idx, reg in enumerate(nuevos):
        fila = fi + idx
        bg   = "0D1B2A" if fila % 2 == 0 else "111827"
        fill = PatternFill("solid", fgColor=bg)
        for ci, (_, k) in enumerate(COLS_XLSX, 1):
            c = ws.cell(row=fila, column=ci, value=reg.get(k, ""))
            c.font      = Font(size=10, name="Calibri", color="C8D8E8")
            c.fill      = fill
            c.alignment = Alignment(
                horizontal="left" if ci <= 2 else "center",
                vertical="center")
            c.border = thin()
    wb.save(ruta)

# ─────────────────────────────────────────────────────────────
# HILO DE CÁMARA
# ─────────────────────────────────────────────────────────────
class HiloCamara(threading.Thread):

    def __init__(self, cola_datos, cola_frames, factor_t, cam_idx):

        super().__init__(daemon=True)

        self.cola_datos  = cola_datos

        self.cola_frames = cola_frames

        self.factor_t    = factor_t

        self.cam_idx     = cam_idx

        self.activo      = True



    def stop(self): self.activo = False



    def run(self):

        mp_pose = mp.solutions.pose

        pose    = mp_pose.Pose(model_complexity=1,

                               min_detection_confidence=0.65,

                               min_tracking_confidence=0.55)

        mp_draw = mp.solutions.drawing_utils

        cap = cv2.VideoCapture(self.cam_idx, cv2.CAP_DSHOW)

        if not cap.isOpened():

            self.cola_datos.put({"error": f"No se pudo abrir cámara {self.cam_idx}."}); return

        cap.set(cv2.CAP_PROP_FRAME_WIDTH,  1280)

        cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 720)

        buffer, inicio = [], time.time()



        while self.activo and cap.isOpened():

            ret, frame = cap.read()

            if not ret: time.sleep(0.02); continue

            frame = cv2.flip(frame, 1)

            h, w, _ = frame.shape

            results = pose.process(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))

            perfil  = False



            if results.pose_landmarks:

                lm = results.pose_landmarks.landmark

                dist_h = abs(lm[mp_pose.PoseLandmark.LEFT_SHOULDER].x -

                             lm[mp_pose.PoseLandmark.RIGHT_SHOULDER].x) * w

                if dist_h < 130:

                    perfil = True

                    lado = ("LEFT"

                            if lm[mp_pose.PoseLandmark.LEFT_SHOULDER].visibility >

                               lm[mp_pose.PoseLandmark.RIGHT_SHOULDER].visibility

                            else "RIGHT")

                    def gp(n):

                        idx = getattr(mp_pose.PoseLandmark, f"{lado}_{n}")

                        return [lm[idx].x * w, lm[idx].y * h]

                    try:

                        ear  = gp("EAR");    sh  = gp("SHOULDER")

                        hip  = gp("HIP");    kn  = gp("KNEE")

                        ank  = gp("ANKLE");  el  = gp("ELBOW")

                        wr   = gp("WRIST")

                        vert = [hip[0], hip[1] + 200]

                        ang_tronco  = calcular_angulo(sh, hip, vert)

                        desv_cuello = abs(180 - calcular_angulo(ear, sh, hip))

                        ang_rodilla = calcular_angulo(hip, kn, ank)

                        ang_codo    = calcular_angulo(sh, el, wr)

                        buffer.append([ang_tronco, desv_cuello, ang_rodilla, ang_codo])

                        mp_draw.draw_landmarks(

                            frame, results.pose_landmarks,

                            mp_pose.POSE_CONNECTIONS,

                            mp_draw.DrawingSpec(color=(0,200,120), thickness=2, circle_radius=3),

                            mp_draw.DrawingSpec(color=(0,160,255), thickness=2))

                        for i, (lbl, val) in enumerate([

                                ("Tronco",  ang_tronco), ("Cuello",  desv_cuello),

                                ("Rodilla", ang_rodilla),("Codo",    ang_codo)]):

                            cv2.putText(frame, f"{lbl}: {val:.0f}deg",

                                        (12, h-14-i*18),

                                        cv2.FONT_HERSHEY_SIMPLEX, 0.44, (0,200,120), 1)

                    except: pass



            cv2.rectangle(frame, (0,0), (w,52), (10,12,18), -1)

            txt   = "PERFIL DETECTADO" if perfil else "POSICIONATE DE PERFIL"

            color = (0,220,80) if perfil else (0,140,255)

            cv2.putText(frame, txt, (14,34), cv2.FONT_HERSHEY_SIMPLEX, 0.72, color, 2)

            restante = max(0, INTERVALO_SEG - int(time.time() - inicio))

            cv2.putText(frame, f"EVAL EN {restante}s", (w-130,34),

                        cv2.FONT_HERSHEY_SIMPLEX, 0.55, (80,160,255), 1)



            if not self.cola_frames.full():

                self.cola_frames.put(cv2.cvtColor(frame, cv2.COLOR_BGR2RGB))



            if time.time() - inicio >= INTERVALO_SEG and len(buffer) >= 3:

                avg = np.mean(np.array(buffer), axis=0)

                rosa, ss, sd, det = calcular_ROSA(

                    float(avg[0]), float(avg[1]), float(avg[2]), float(avg[3]),

                    self.factor_t)

                if not self.cola_datos.full():

                    self.cola_datos.put({

                        "rosa": rosa, "score_silla": ss, "score_D": sd, "detalle": det,

                        "angulos": {

                            "ang_tronco_deg":      round(float(avg[0]), 1),

                            "ang_cuello_desv_deg": round(float(avg[1]), 1),

                            "ang_rodilla_deg":     round(float(avg[2]), 1),

                            "ang_codo_deg":        round(float(avg[3]), 1),

                        }})

                buffer, inicio = [], time.time()



        cap.release()

        self.cola_datos.put({"fin": True})

# ─────────────────────────────────────────────────────────────
# VENTANA DE ALERTA — solo score >= 5 (NTP 1173)
# Cierre: mover el ratón
# ─────────────────────────────────────────────────────────────
class VentanaAlerta(tk.Toplevel):
    def __init__(self, parent, rosa, nombre, callback_cerrar=None):
        super().__init__(parent)
        self._activa   = True
        self._callback = callback_cerrar
        self._blink_id = None
        self._construir(rosa, nombre)
        disparar_sonido(veces=4, freq=1200, dur=320)
        self.bind("<Motion>",   lambda e: self.cerrar())
        self.bind("<Button-1>", lambda e: self.cerrar())
        self.bind("<Key>",      lambda e: self.cerrar())

    def _construir(self, rosa, nombre):
        bg = "#8B0000" if rosa >= 8 else (C["danger2"] if rosa >= 6 else C["danger"])
        self._bg = bg
        self.overrideredirect(True)
        self.attributes("-topmost", True)
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"{sw}x{sh}+0+0")
        self.configure(bg=bg)

        tk.Label(self, text=str(rosa), bg=bg, fg="white",
                 font=("Courier New", 200, "bold")
                 ).place(relx=0.5, rely=0.27, anchor="center")

        cv_ = tk.Canvas(self, bg=bg, highlightthickness=0, height=3, width=sw-120)
        cv_.place(relx=0.5, rely=0.52, anchor="center")
        cv_.create_line(0, 1, sw-120, 1, fill="white", width=2)

        tk.Label(self, text="NIVEL DE ACCION  ·  NTP 1173", bg=bg, fg="white",
                 font=("Courier New", 40, "bold")
                 ).place(relx=0.5, rely=0.60, anchor="center")

        tk.Label(self,
                 text="Puntuacion ROSA >= 5  —  Se requieren cambios inmediatos en el puesto de trabajo",
                 bg=bg, fg="#FFCCCC", font=("Courier New", 17)
                 ).place(relx=0.5, rely=0.70, anchor="center")

        tk.Label(self, text=f"Trabajador: {nombre}", bg=bg, fg="white",
                 font=("Courier New", 15)
                 ).place(relx=0.5, rely=0.79, anchor="center")

        tk.Label(self, text="Mueve el raton para cerrar", bg=bg, fg="#FF9999",
                 font=("Courier New", 12)
                 ).place(relx=0.5, rely=0.93, anchor="center")

        self._blink(bg)

    def _blink(self, base):
        if not self._activa: return
        c   = self.cget("bg")
        alt = "#5C0000" if c == base else base
        try:
            self.configure(bg=alt)
            for w in self.winfo_children():
                try: w.configure(bg=alt)
                except: pass
        except: return
        self._blink_id = self.after(700, lambda: self._blink(base))

    def cerrar(self):
        if not self._activa: return
        self._activa = False
        if self._blink_id:
            try: self.after_cancel(self._blink_id)
            except: pass
        try: self.destroy()
        except: pass
        if self._callback: self._callback()

# ─────────────────────────────────────────────────────────────
# PANTALLA DE INICIO
# ─────────────────────────────────────────────────────────────
def _style_combo_popup(cb):
    try:
        cb.tk.eval(f"""
            set popdown [ttk::combobox::PopdownWindow {cb}]
            $popdown.f.l configure \
                -background {C["bg_card"]} \
                -foreground {C["text_hi"]} \
                -selectbackground {C["accent2"]} \
                -selectforeground {C["text_hi"]} \
                -font {{"Courier New" 11}}
        """)
    except Exception:
        pass

class PantallaInicio(tk.Toplevel):
    W, H = 500, 600

    def __init__(self, parent, camaras):
        super().__init__(parent)
        self.title("ROSA NTP 1173")
        self.resizable(False, False)
        self.grab_set()
        self.resultado = None; self.camaras = camaras
        self.configure(bg=C["bg_deep"])
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"{self.W}x{self.H}+{(sw-self.W)//2}+{(sh-self.H)//2}")
        self._build()
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.wait_window()

    def _mk_field(self, parent, placeholder=""):
        wrap = tk.Frame(parent, bg=C["bg_card"],
                        highlightthickness=1, highlightbackground=C["bg_border"])
        wrap.pack(fill="x", padx=44, pady=(0,0))
        e = tk.Entry(wrap, bg=C["bg_card"], fg=C["text_hi"],
                     insertbackground=C["accent"], relief="flat", bd=0,
                     font=("Courier New", 12), highlightthickness=0)
        e.pack(fill="x", padx=12, pady=9)
        return e, wrap

    def _mk_combo(self, parent, values):
        sty = ttk.Style(); sty.theme_use("clam")
        sty.configure("D.TCombobox",
            fieldbackground=C["bg_card"], background=C["bg_card"],
            foreground=C["text_hi"], insertcolor=C["text_hi"],
            arrowcolor=C["accent"], bordercolor=C["bg_border"],
            lightcolor=C["bg_card"], darkcolor=C["bg_card"],
            selectforeground=C["text_hi"], selectbackground=C["accent2"],
            padding=(8,6))
        sty.map("D.TCombobox",
            fieldbackground=[("readonly",C["bg_card"]),("disabled",C["bg_deep"])],
            foreground=[("readonly",C["text_hi"]),("disabled",C["text_lo"])],
            selectforeground=[("readonly",C["text_hi"])],
            selectbackground=[("readonly",C["accent2"])],
            background=[("active",C["bg_border"]),("pressed",C["bg_border"])])
        wrap = tk.Frame(parent, bg=C["bg_card"],
                        highlightthickness=1, highlightbackground=C["bg_border"])
        wrap.pack(fill="x", padx=44)
        cb = ttk.Combobox(wrap, values=values, state="readonly",
                          style="D.TCombobox", font=("Courier New", 11))
        cb.pack(fill="x", padx=6, pady=4)
        cb.bind("<Map>", lambda e, c=cb: _style_combo_popup(c))
        return cb

    def _lbl_section(self, parent, txt):
        tk.Label(parent, text=txt, bg=C["bg_deep"], fg=C["text_mid"],
                 font=("Courier New", 8, "bold"), anchor="w"
                 ).pack(fill="x", padx=44, pady=(14,3))

    def _build(self):
        hdr = tk.Frame(self, bg=C["bg_panel"], height=108)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr, text="SISTEMA AUTOMATIZADO EVALUACIÓN ROSA",
                 bg=C["bg_panel"], fg=C["accent"],
                 font=("Courier New", 16, "bold")
                 ).place(relx=0.5, rely=0.30, anchor="center")
        tk.Label(hdr, text="PROYECTO TESIS  ·  JAIME TARAZONA",
                 bg=C["bg_panel"], fg=C["text_mid"],
                 font=("Courier New", 12, "bold")
                 ).place(relx=0.5, rely=0.55, anchor="center")

        tk.Frame(self, bg=C["accent2"], height=2).pack(fill="x")

        self._lbl_section(self, "NOMBRE DEL TRABAJADOR")
        self.e_nombre, self._wrap_nombre = self._mk_field(self)
        self.e_nombre.focus()

        self._lbl_section(self, "ID / EXPEDIENTE")
        self.e_id, self._wrap_id = self._mk_field(self)

        self._lbl_section(self, "USO DIARIO  (Tabla F · NTP 1173)")
        self.cb_uso = self._mk_combo(self, [
            "< 1 hora / dia       —  factor  -1",
            "1 - 4 horas / dia    —  factor   0",
            "> 4 horas / dia      —  factor  +1",
        ])
        self.cb_uso.current(1)

        self._lbl_section(self, "CAMARA")
        self.cb_cam = self._mk_combo(self, [lbl for _, lbl in self.camaras])
        self.cb_cam.current(0)

        if len(self.camaras) == 1:
            tk.Label(self, text="   Solo se detecto una camara",
                     bg=C["bg_deep"], fg=C["text_mid"],
                     font=("Courier New", 8)).pack(anchor="w", padx=44)

        tk.Frame(self, bg=C["bg_deep"]).pack(fill="both", expand=True)
        tk.Frame(self, bg=C["bg_border"], height=1).pack(fill="x")

        foot = tk.Frame(self, bg=C["bg_panel"], height=72)
        foot.pack(fill="x"); foot.pack_propagate(False)
        self.btn = tk.Button(
            foot, text="  \u25cf  INICIAR GRABACION",
            bg=C["danger"], fg="white",
            activebackground=C["danger2"], activeforeground="white",
            font=("Courier New", 14, "bold"),
            relief="flat", bd=0, cursor="hand2",
            command=self._ok)
        self.btn.place(relx=0.5, rely=0.5, anchor="center", width=340, height=46)
        self.btn.bind("<Enter>", lambda e: self.btn.configure(bg="#FF2D3D"))
        self.btn.bind("<Leave>", lambda e: self.btn.configure(bg=C["danger"]))

        self.bind("<Return>", lambda _: self._ok())
        self.bind("<Escape>", lambda _: self.destroy())

    def _ok(self):
        nombre = self.e_nombre.get().strip()
        id_t   = self.e_id.get().strip()
        err    = False
        for e, w, v in [(self.e_nombre, self._wrap_nombre, nombre),
                        (self.e_id,     self._wrap_id,     id_t)]:
            if not v:
                w.configure(highlightbackground=C["danger"], highlightcolor=C["danger"])
                err = True
            else:
                w.configure(highlightbackground=C["bg_border"], highlightcolor=C["accent"])
        if err: return
        horas   = {0: 0.5, 1: 2.0, 2: 5.0}[self.cb_uso.current()]
        cam_idx = self.camaras[self.cb_cam.current()][0]
        self.resultado = {"nombre": nombre, "id": id_t, "horas": horas, "cam_idx": cam_idx}
        self.destroy()

# ─────────────────────────────────────────────────────────────
# PANEL PRINCIPAL
# ─────────────────────────────────────────────────────────────
class PanelROSA(tk.Tk):
    CAM_W, CAM_H = 660, 495
    PANEL_W      = 400

    def __init__(self, nombre, id_t, uso_horas, cam_idx):
        super().__init__()
        self.nombre      = nombre; self.id_t        = id_t
        self.uso_horas   = uso_horas; self.cam_idx  = cam_idx
        self.factor_t    = factor_tiempo_F(uso_horas)
        self.inicio_sesion = time.strftime("%Y%m%d_%H%M%S")
        self.registros   = []; self._reg_guardados  = 0
        self.toaster     = ToastNotifier() if TOAST_OK else None
        self._ventana_alerta   = None
        self._ultimo_resultado = None
        self.cola_datos  = queue.Queue(maxsize=5)
        self.cola_frames = queue.Queue(maxsize=2)
        self._build()
        self._iniciar_camara()
        self._tick_cd()
        self._guardado_periodico()
        self._poll()

    def _build(self):
        self.title(f"ROSA NTP 1173  ·  {self.nombre}  ·  {self.id_t}")
        self.configure(bg=C["bg_deep"])
        W = self.CAM_W + self.PANEL_W + 28
        H = self.CAM_H + 78
        self.geometry(f"{W}x{H}")
        self.resizable(False, False)

        sty = ttk.Style(); sty.theme_use("clam")
        sty.configure("Treeview",
            background=C["bg_card"], fieldbackground=C["bg_card"],
            foreground=C["text_mid"], rowheight=20, font=("Courier New",9))
        sty.configure("Treeview.Heading",
            background=C["bg_panel"], foreground=C["accent"],
            font=("Courier New",9,"bold"), relief="flat")
        sty.map("Treeview",
            background=[("selected",C["bg_border"])],
            foreground=[("selected",C["text_hi"])])
        sty.configure("Bar.Horizontal.TProgressbar",
            troughcolor=C["bg_panel"], background=C["ok"],
            thickness=6, bordercolor=C["bg_panel"])

        self.columnconfigure(0, minsize=self.CAM_W+12)
        self.columnconfigure(1, minsize=self.PANEL_W)
        self.rowconfigure(0, weight=1)

        self._build_camara()
        self._build_panel()
        self._build_status()
        self.protocol("WM_DELETE_WINDOW", self._detener)

    def _build_camara(self):
        frm = tk.Frame(self, bg=C["bg_panel"],
                       width=self.CAM_W+12, height=self.CAM_H+52)
        frm.grid(row=0, column=0, sticky="nsew", padx=(8,4), pady=(8,0))
        frm.pack_propagate(False); frm.grid_propagate(False)

        hdr = tk.Frame(frm, bg=C["bg_panel"], height=28)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr, text="  FEED DE CAMARA",
                 bg=C["bg_panel"], fg=C["accent"],
                 font=("Courier New",9,"bold")).pack(side="left", pady=5)
        self.lbl_cam_estado = tk.Label(hdr, text="INICIANDO",
                 bg=C["bg_panel"], fg=C["text_lo"], font=("Courier New",8))
        self.lbl_cam_estado.pack(side="right", padx=10)

        tk.Frame(frm, bg=C["bg_border"], height=1).pack(fill="x")

        self.cam_canvas = tk.Canvas(frm, width=self.CAM_W, height=self.CAM_H,
                                    bg="#000", highlightthickness=0)
        self.cam_canvas.pack(padx=6, pady=6)
        self._img_id = self.cam_canvas.create_image(0, 0, anchor="nw")

        cd_row = tk.Frame(frm, bg=C["bg_panel"])
        cd_row.pack(fill="x", padx=6, pady=(0,6))
        tk.Label(cd_row, text="PROX EVAL",
                 bg=C["bg_panel"], fg=C["text_lo"],
                 font=("Courier New",7)).pack(side="left", padx=(4,6))
        self.barra_cd = ttk.Progressbar(cd_row, style="Bar.Horizontal.TProgressbar",
                                         maximum=INTERVALO_SEG, length=self.CAM_W-110)
        self.barra_cd.pack(side="left")
        self.lbl_cd = tk.Label(cd_row, text=f"{INTERVALO_SEG}s",
                               bg=C["bg_panel"], fg=C["accent"],
                               font=("Courier New",9,"bold"), width=5)
        self.lbl_cd.pack(side="left", padx=6)

    def _build_panel(self):
        frm = tk.Frame(self, bg=C["bg_deep"],
                       width=self.PANEL_W, height=self.CAM_H+52)
        frm.grid(row=0, column=1, sticky="nsew", padx=(4,8), pady=(8,0))
        frm.pack_propagate(False); frm.grid_propagate(False)

        cs = tk.Frame(frm, bg=C["bg_card"],
                      highlightthickness=1, highlightbackground=C["bg_border"])
        cs.pack(fill="x", pady=(0,5))

        row_s = tk.Frame(cs, bg=C["bg_card"])
        row_s.pack(fill="x", padx=10, pady=(10,2))

        self.lbl_score = tk.Label(row_s, text="—", bg=C["bg_card"], fg=C["accent"],
                                  font=("Courier New",78,"bold"), width=3, anchor="center")
        self.lbl_score.pack(side="left")

        meta = tk.Frame(row_s, bg=C["bg_card"])
        meta.pack(side="left", fill="y", padx=(6,0), pady=4)

        tk.Label(meta, text="PUNTUACION ROSA",
                 bg=C["bg_card"], fg=C["text_lo"],
                 font=("Courier New",8,"bold")).pack(anchor="w")
        self.lbl_nivel = tk.Label(meta, text="Esperando...",
                                  bg=C["bg_card"], fg=C["text_mid"],
                                  font=("Courier New",12,"bold"))
        self.lbl_nivel.pack(anchor="w", pady=(2,8))

        for attr, lbl_txt in [("lbl_silla", "SILLA  (A + F)"),
                               ("lbl_perif", "PERIFERICOS  (D)")]:
            tk.Label(meta, text=lbl_txt, bg=C["bg_card"], fg=C["text_lo"],
                     font=("Courier New",8)).pack(anchor="w")
            lbl = tk.Label(meta, text="—", bg=C["bg_card"], fg=C["text_hi"],
                           font=("Courier New",11,"bold"))
            lbl.pack(anchor="w")
            setattr(self, attr, lbl)

        marks = tk.Frame(cs, bg=C["bg_card"])
        marks.pack(fill="x", padx=10, pady=(6,2))
        for i in range(1, 11):
            mc = C["danger"] if i >= 5 else (C["warn"] if i >= 3 else C["ok"])
            tk.Label(marks, text=str(i), bg=C["bg_card"], fg=mc,
                     font=("Courier New",7), width=3).pack(side="left", expand=True)

        self.barra_rosa = ttk.Progressbar(cs, style="Bar.Horizontal.TProgressbar",
                                           maximum=10, length=self.PANEL_W-20)
        self.barra_rosa.pack(fill="x", padx=10, pady=(0,4))

        tk.Label(cs, text="NTP 1173  ·  Escala 1-10  ·  Nivel de accion >= 5",
                 bg=C["bg_card"], fg=C["text_lo"],
                 font=("Courier New",7)).pack(anchor="w", padx=10, pady=(0,8))

        cd2 = tk.Frame(frm, bg=C["bg_card"],
                       highlightthickness=1, highlightbackground=C["bg_border"])
        cd2.pack(fill="both", expand=True)

        tk.Label(cd2, text="  DESGLOSE  TABLAS NTP 1173",
                 bg=C["bg_card"], fg=C["accent"],
                 font=("Courier New",8,"bold")).pack(anchor="w", pady=(8,4))
        tk.Frame(cd2, bg=C["bg_border"], height=1).pack(fill="x")

        tree_frm = tk.Frame(cd2, bg=C["bg_card"])
        tree_frm.pack(fill="both", expand=True, padx=0)

        self.tree = ttk.Treeview(tree_frm, columns=("F","V"),
                                  show="headings", selectmode="none")
        self.tree.heading("F", text="Factor NTP 1173")
        self.tree.heading("V", text="Punt.")
        self.tree.column("F", width=272, anchor="w")
        self.tree.column("V", width=50,  anchor="center")
        sb = ttk.Scrollbar(tree_frm, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side="left", fill="both", expand=True, padx=(4,0))
        sb.pack(side="right", fill="y", pady=2)
        self.tree.tag_configure("critico", foreground=C["danger"])
        self.tree.tag_configure("alto",    foreground=C["warn"])
        self.tree.tag_configure("ok",      foreground=C["ok"])
        self.tree.tag_configure("neutro",  foreground=C["text_mid"])
        self.tree.tag_configure("final",   foreground=C["accent"],
                                            font=("Courier New",9,"bold"))

    def _build_status(self):
        bar = tk.Frame(self, bg=C["bg_panel"], height=28)
        bar.grid(row=1, column=0, columnspan=2, sticky="ew")
        bar.pack_propagate(False); bar.grid_propagate(False)
        tk.Frame(bar, bg=C["bg_border"], height=1).pack(fill="x", side="top")
        tk.Label(bar,
                 text=f"  {self.nombre.upper()}   |   {self.id_t}   |   {self.uso_horas}h/dia   |   factor F = {self.factor_t:+d}",
                 bg=C["bg_panel"], fg=C["text_lo"],
                 font=("Courier New",8)).pack(side="left", pady=5)
        self.lbl_guardado = tk.Label(bar, text="AUTO-SAVE: —",
                 bg=C["bg_panel"], fg=C["text_lo"], font=("Courier New",8))
        self.lbl_guardado.pack(side="right", padx=14)

    def _tick_cd(self):
        if not hasattr(self, "_cd_val"): self._cd_val = INTERVALO_SEG
        self._cd_val = max(0, self._cd_val - 1)
        try:
            self.lbl_cd.configure(text=f"{self._cd_val}s")
            self.barra_cd["value"] = INTERVALO_SEG - self._cd_val
        except: pass
        self.after(1000, self._tick_cd)

    def _guardado_periodico(self):
        try:
            if self._ultimo_resultado is not None:
                r  = self._ultimo_resultado
                ts = time.strftime("%Y-%m-%d %H:%M:%S")
                nivel_txt, _ = nivel_accion(r["rosa"])
                self.registros.append({
                    "id_evaluado":         self.id_t,
                    "nombre_evaluado":     self.nombre,
                    "horas_uso_diario":    self.uso_horas,
                    "timestamp":           ts,
                    "fecha":               ts[:10],
                    "hora":                ts[11:],
                    "rosa_final":          r["rosa"],
                    "nivel_riesgo":        nivel_txt,
                    "alerta_disparada":    "SI" if r["rosa"] >= NIVEL_ACCION else "NO",
                    "score_silla":         r["ss"],
                    "score_perifericos_D": r["sd"],
                    "p_A1_altura":         r["detalle"].get("p_A1_altura",    ""),
                    "p_A2_profundidad":    r["detalle"].get("p_A2_profundidad",""),
                    "p_A3_reposabrazos":   r["detalle"].get("p_A3_reposabrazos",""),
                    "p_A4_respaldo":       r["detalle"].get("p_A4_respaldo",   ""),
                    "tabla_A":             r["detalle"].get("tabla_A",         ""),
                    "factor_F":            r["detalle"].get("factor_F",        ""),
                    "p_B1_telefono":       r["detalle"].get("p_B1_telefono",   ""),
                    "p_B2_pantalla":       r["detalle"].get("p_B2_pantalla",   ""),
                    "tabla_B":             r["detalle"].get("tabla_B",         ""),
                    "p_C1_raton":          r["detalle"].get("p_C1_raton",      ""),
                    "p_C2_teclado":        r["detalle"].get("p_C2_teclado",    ""),
                    "tabla_C":             r["detalle"].get("tabla_C",         ""),
                    "tabla_D":             r["detalle"].get("tabla_D",         ""),
                    "ang_tronco_deg":      r["angulos"].get("ang_tronco_deg",      ""),
                    "ang_cuello_desv_deg": r["angulos"].get("ang_cuello_desv_deg", ""),
                    "ang_rodilla_deg":     r["angulos"].get("ang_rodilla_deg",     ""),
                    "ang_codo_deg":        r["angulos"].get("ang_codo_deg",        ""),
                })
                exportar_xlsx(self._ruta_xlsx(), self.registros)
                self._reg_guardados = len(self.registros)
                try:
                    self.lbl_guardado.configure(
                        text=f"AUTO-SAVE: {ts[11:]}  ({len(self.registros)} reg.)",
                        fg=C["ok"])
                except: pass
        except Exception as e:
            print(f"[ROSA] Guardado auto: {e}")
        self.after(GUARDADO_SEG * 1000, self._guardado_periodico)

    def _iniciar_camara(self):
        self.hilo = HiloCamara(self.cola_datos, self.cola_frames,
                               self.factor_t, self.cam_idx)
        self.hilo.start()

    def _poll(self):
        try:
            rgb = self.cola_frames.get_nowait()
            img = Image.fromarray(rgb).resize((self.CAM_W, self.CAM_H), Image.LANCZOS)
            imgtk = ImageTk.PhotoImage(image=img)
            self.cam_canvas.itemconfig(self._img_id, image=imgtk)
            self.cam_canvas._imgtk = imgtk
        except queue.Empty: pass

        try:
            datos = self.cola_datos.get_nowait()
            if "error" in datos:
                messagebox.showerror("Error de camara", datos["error"])
                self._detener(); return
            if "fin" in datos:
                self._detener(); return

            rosa    = datos["rosa"]; ss     = datos["score_silla"]
            sd      = datos["score_D"]; detalle = datos["detalle"]
            angulos = datos.get("angulos", {})
            nivel_txt, col = nivel_accion(rosa)
            self._cd_val = INTERVALO_SEG

            self.lbl_score.configure(text=str(rosa), fg=col)
            self.lbl_nivel.configure(text=nivel_txt,  fg=col)
            self.lbl_silla.configure(text=str(ss))
            self.lbl_perif.configure(text=str(sd))
            self.barra_rosa["value"] = rosa
            ttk.Style().configure("Bar.Horizontal.TProgressbar", background=col)
            try:
                self.lbl_cam_estado.configure(
                    text=f"ROSA={rosa}  S={ss}  D={sd}", fg=col)
            except: pass

            ETIQ = {
                "p_A1_altura":       "A-1  Altura asiento",
                "p_A2_profundidad":  "A-2  Profundidad",
                "p_A3_reposabrazos": "A-3  Reposabrazos",
                "p_A4_respaldo":     "A-4  Respaldo",
                "tabla_A":           "Tabla A  (silla)",
                "factor_F":          "Tabla F  (tiempo uso)",
                "score_silla":       "Score Silla  (A + F)",
                "p_B1_telefono":     "B-1  Telefono",
                "p_B2_pantalla":     "B-2  Pantalla",
                "tabla_B":           "Tabla B  (tel + pant)",
                "p_C1_raton":        "C-1  Raton",
                "p_C2_teclado":      "C-2  Teclado",
                "tabla_C":           "Tabla C  (rat + tec)",
                "tabla_D":           "Tabla D  (perifericos)",
                "rosa_final":        "> ROSA FINAL",
            }
            for item in self.tree.get_children(): self.tree.delete(item)
            for k, v in detalle.items():
                if k == "rosa_final":               tag = "final"
                elif isinstance(v, int) and v >= 5: tag = "critico"
                elif isinstance(v, int) and v >= 3: tag = "alto"
                elif isinstance(v, int) and v >= 2: tag = "neutro"
                else:                               tag = "ok"
                self.tree.insert("", "end", values=(ETIQ.get(k, k), v), tags=(tag,))

            self._gestionar_alerta(rosa)
            self._ultimo_resultado = {
                "rosa": rosa, "ss": ss, "sd": sd,
                "detalle": detalle, "angulos": angulos,
            }

        except queue.Empty: pass
        self.after(40, self._poll)

    def _gestionar_alerta(self, rosa):
        if rosa >= NIVEL_ACCION:
            if self._ventana_alerta is None:
                self._ventana_alerta = VentanaAlerta(
                    self, rosa, self.nombre,
                    callback_cerrar=lambda: setattr(self, "_ventana_alerta", None))
                disparar_toast(self.toaster,
                               f"ROSA {rosa}/10  -  Nivel de Accion",
                               f"{self.nombre}\nActuacion inmediata requerida (NTP 1173)")
        else:
            if self._ventana_alerta:
                try: self._ventana_alerta.cerrar()
                except: pass
                self._ventana_alerta = None

    def _ruta_xlsx(self):
        base = os.path.dirname(os.path.abspath(
            sys.executable if getattr(sys, "frozen", False) else __file__))
        return os.path.join(base, f"ROSA_{self.id_t}_{self.inicio_sesion}.xlsx")

    def _detener(self):
        if self._ventana_alerta:
            try: self._ventana_alerta.cerrar()
            except: pass
        if hasattr(self, "hilo"): self.hilo.stop()
        if self._ultimo_resultado is not None:
            try:
                r  = self._ultimo_resultado
                ts = time.strftime("%Y-%m-%d %H:%M:%S")
                nivel_txt, _ = nivel_accion(r["rosa"])
                self.registros.append({
                    "id_evaluado":       self.id_t,
                    "nombre_evaluado":   self.nombre,
                    "horas_uso_diario":  self.uso_horas,
                    "timestamp": ts, "fecha": ts[:10], "hora": ts[11:],
                    "rosa_final":        r["rosa"],
                    "nivel_riesgo":      nivel_txt,
                    "alerta_disparada":  "SI" if r["rosa"] >= NIVEL_ACCION else "NO",
                    "score_silla":         r["ss"],
                    "score_perifericos_D": r["sd"],
                    "p_A1_altura":         r["detalle"].get("p_A1_altura",    ""),
                    "p_A2_profundidad":    r["detalle"].get("p_A2_profundidad",""),
                    "p_A3_reposabrazos":   r["detalle"].get("p_A3_reposabrazos",""),
                    "p_A4_respaldo":       r["detalle"].get("p_A4_respaldo",   ""),
                    "tabla_A":             r["detalle"].get("tabla_A",         ""),
                    "factor_F":            r["detalle"].get("factor_F",        ""),
                    "p_B1_telefono":       r["detalle"].get("p_B1_telefono",   ""),
                    "p_B2_pantalla":       r["detalle"].get("p_B2_pantalla",   ""),
                    "tabla_B":             r["detalle"].get("tabla_B",         ""),
                    "p_C1_raton":          r["detalle"].get("p_C1_raton",      ""),
                    "p_C2_teclado":        r["detalle"].get("p_C2_teclado",    ""),
                    "tabla_C":             r["detalle"].get("tabla_C",         ""),
                    "tabla_D":             r["detalle"].get("tabla_D",         ""),
                    "ang_tronco_deg":      r["angulos"].get("ang_tronco_deg",      ""),
                    "ang_cuello_desv_deg": r["angulos"].get("ang_cuello_desv_deg", ""),
                    "ang_rodilla_deg":     r["angulos"].get("ang_rodilla_deg",     ""),
                    "ang_codo_deg":        r["angulos"].get("ang_codo_deg",        ""),
                })
            except: pass
        if self.registros:
            ruta = self._ruta_xlsx()
            try:
                exportar_xlsx(ruta, self.registros)
                messagebox.showinfo("Sesion guardada",
                                    f"Informe guardado:\n{ruta}\n{len(self.registros)} registros.",
                                    parent=self)
                try: os.startfile(ruta)
                except: pass
            except Exception as e:
                messagebox.showerror("Error al guardar", str(e), parent=self)
        self.after(300, self.destroy)

# ─────────────────────────────────────────────────────────────
# PUNTO DE ENTRADA
# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    raiz = tk.Tk(); raiz.withdraw()
    camaras = []
    def _scan(): camaras.extend(detectar_camaras())
    t = threading.Thread(target=_scan, daemon=True); t.start(); t.join(timeout=10)
    if not camaras: camaras = [(0, "Camara 0 (por defecto)")]
    dialogo = PantallaInicio(raiz, camaras)
    datos   = dialogo.resultado
    raiz.destroy()
    if not datos: sys.exit(0)
    PanelROSA(
        nombre    = datos["nombre"],
        id_t      = datos["id"],
        uso_horas = datos["horas"],
        cam_idx   = datos["cam_idx"],
    ).mainloop()